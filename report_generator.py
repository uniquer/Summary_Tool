"""
Excel report generation module
"""
import pandas as pd
from datetime import datetime
from typing import List, Dict, Optional
import os


class ReportGenerator:
    """Generates Excel reports from summary data"""

    def __init__(self):
        """Initialize report generator"""
        pass

    def create_excel_report(self, summaries: List[Dict], output_filename: Optional[str] = None) -> str:
        """
        Create Excel report from summary data

        Args:
            summaries: List of summary dictionaries
            output_filename: Optional custom filename

        Returns:
            Path to generated Excel file
        """
        # Prepare data for DataFrame
        report_data = []

        for summary in summaries:
            report_data.append({
                'URL': summary.get('url', ''),
                'Filename': summary.get('filename', ''),
                'Status': summary.get('status', ''),
                'Long Summary': summary.get('long_summary', ''),
                'Short Summary': summary.get('short_summary', ''),
                'Error Message': summary.get('error_message', ''),
                'Created At': summary.get('created_at', ''),
            })

        # Create DataFrame
        df = pd.DataFrame(report_data)

        # Generate filename
        if not output_filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f'summary_report_{timestamp}.xlsx'

        # Ensure .xlsx extension
        if not output_filename.endswith('.xlsx'):
            output_filename += '.xlsx'

        # Create Excel writer with formatting
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Summaries', index=False)

            # Get the worksheet
            worksheet = writer.sheets['Summaries']

            # Set column widths
            column_widths = {
                'A': 50,  # URL
                'B': 30,  # Filename
                'C': 15,  # Status
                'D': 80,  # Long Summary
                'E': 60,  # Short Summary
                'F': 40,  # Error Message
                'G': 20,  # Created At
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # Enable text wrapping for summary columns
            from openpyxl.styles import Alignment

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Make header bold
            from openpyxl.styles import Font

            for cell in worksheet[1]:
                cell.font = Font(bold=True)

            # Add auto-filter
            worksheet.auto_filter.ref = worksheet.dimensions

        return output_filename

    def create_report_from_session(self, session_data: List[Dict]) -> str:
        """
        Create report from current session data

        Args:
            session_data: List of processed items from current session

        Returns:
            Path to generated Excel file
        """
        return self.create_excel_report(session_data)

    def create_download_report(self, results: List[Dict], output_filename: Optional[str] = None) -> str:
        """
        Create Excel report for downloads (Tab 1)
        Columns: Link, File Name, Download Status/Error
        """
        report_data = []
        for result in results:
            status = result.get('download_status', 'pending')
            error = result.get('download_error', '')
            status_msg = status
            if status == 'failed' and error:
                status_msg = f"Failed: {error}"
            elif status == 'skipped':
                status_msg = "Already Exists"
            
            report_data.append({
                'Link': result.get('url', ''),
                'File Name': result.get('filename', ''),
                'Download Status': status_msg
            })
            
        df = pd.DataFrame(report_data)
        
        if not output_filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f'download_report_{timestamp}.xlsx'
            
        if not output_filename.endswith('.xlsx'):
            output_filename += '.xlsx'
            
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Downloads', index=False)
            worksheet = writer.sheets['Downloads']
            
            # Formatting
            worksheet.column_dimensions['A'].width = 60
            worksheet.column_dimensions['B'].width = 40
            worksheet.column_dimensions['C'].width = 50
            
            from openpyxl.styles import Alignment, Font
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                
        return output_filename

    def create_summary_report(self, summaries: List[Dict], download_folder: str, output_filename: Optional[str] = None) -> str:
        """
        Create Excel report for summaries (Tab 2)
        Columns: Link, File Name, Long Summary, Short Summary, Date Downloaded, Date Summarized
        """
        report_data = []
        for summary in summaries:
            filename = summary.get('filename', '')
            
            # Get date downloaded from file system
            date_downloaded = "Unknown"
            if filename and download_folder:
                filepath = os.path.join(download_folder, filename)
                if os.path.exists(filepath):
                    timestamp = os.path.getmtime(filepath)
                    date_downloaded = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
            
            # Get date summarized
            date_summarized = summary.get('created_at', '')
            if summary.get('status') != 'success':
                date_summarized = f"Failed: {summary.get('error_message', 'Unknown error')}"
            
            report_data.append({
                'Link': summary.get('url', ''),
                'File Name': filename,
                'Long Summary': summary.get('long_summary', ''),
                'Short Summary': summary.get('short_summary', ''),
                'Error Message': summary.get('error_message', ''),
                'Date Downloaded': date_downloaded,
                'Date Summarized': date_summarized
            })
            
        df = pd.DataFrame(report_data)
        
        if not output_filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f'summary_report_{timestamp}.xlsx'
            
        if not output_filename.endswith('.xlsx'):
            output_filename += '.xlsx'
            
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Summaries', index=False)
            worksheet = writer.sheets['Summaries']
            
            # Formatting
            column_widths = {'A': 50, 'B': 30, 'C': 80, 'D': 60, 'E': 40, 'F': 20, 'G': 20}
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            from openpyxl.styles import Alignment, Font
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                
        return output_filename

    def get_summary_statistics(self, summaries: List[Dict]) -> Dict:
        """
        Calculate statistics from summaries

        Args:
            summaries: List of summary dictionaries

        Returns:
            Dictionary with statistics
        """
        total = len(summaries)
        successful = sum(1 for s in summaries if s.get('status') == 'success')
        failed = total - successful

        return {
            'total': total,
            'successful': successful,
            'failed': failed,
            'success_rate': (successful / total * 100) if total > 0 else 0
        }
